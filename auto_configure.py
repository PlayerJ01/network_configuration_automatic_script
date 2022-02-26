from asyncio.windows_events import NULL
from turtle import delay
from netmiko import ConnectHandler
from netmiko.ssh_exception import NetMikoTimeoutException
from netmiko.ssh_exception import NetMikoAuthenticationException
import openpyxl
import datetime
import os 

############--执行脚本----##########################################
def dev_configure(ip,type1,port,username,password,secret,cmds,file_path):
    if type1.find('[Rr]uijie') != -1 :
        type1 = 'cisco_ios'
    elif type1.find('[Hh]3[Cc]') != -1   :
        type1 = 'hp_comware'
    device = {'device_type': type1,
          'ip': ip,
          'port': port,
          'username': username,
          'password': password, 
          'secret':secret
          }
    file = file_path + ip + '.txt'
    try:
        with ConnectHandler(**device) as connect:
            connect.enable()
            save = open(file,'a',encoding='utf-8')
            for cmd in cmds:
                output = "---------------" + cmd + "---------------------\n" + connect.send_command_timing(cmd)
                save.write(output)
            save.close()
            connect.disconnect()
    except (EOFError,NetMikoTimeoutException):
        errorfile = file_path + '0_error.txt'
        output = ip + '\t can not connect to Device! \n'
        errorsave = open(errorfile,'a',encoding='utf-8')
        errorsave.write(output)
        errorsave.close()
    except (EOFError, NetMikoAuthenticationException):
        errorfile = file_path + '0_error.txt'
        output = ip + '\t login username/password wrong! \n'
        errorsave = open(errorfile,'a',encoding='utf-8')
        errorsave.write(output)
        errorsave.close()
    except (ValueError, NetMikoAuthenticationException):
        errorfile = file_path + '0_error.txt'
        output = ip + '\t enable password wrong! \n'
        errorsave = open(errorfile,'a',encoding='utf-8')
        errorsave.write(output)
        errorsave.close()


############---主程序----##########################################
now_time = datetime.datetime.now().strftime('%Y_%m_%d')
path = os.getcwd()
source_file = "switch.xlsx"
if not os.path.isdir('./'+now_time):
    os.mkdir(now_time)    
#file_name = "0_test-" + now_time + ".xlsx"
source_wb = openpyxl.load_workbook(source_file)
source_sheet = source_wb.active
row_max = source_sheet.max_row + 1
for i in range(2,row_max):
    dev_ip = source_sheet.cell(row=i,column=1).value
    device_type = source_sheet.cell(row=i,column=2).value
    dev_port=  str(source_sheet.cell(row=i,column=3).value)
    dev_username = source_sheet.cell(row=i,column=4).value
    dev_password = source_sheet.cell(row=i,column=5).value
    dev_secret = source_sheet.cell(row=i,column=6).value
    dev_cmds = source_sheet.cell(row=i,column=7).value.split(';')
    file_path = path + '\\'+now_time+'\\'
    dev_configure(dev_ip,device_type,dev_port,dev_username,dev_password,dev_secret,dev_cmds,file_path)